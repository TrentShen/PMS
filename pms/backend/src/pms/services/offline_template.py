from __future__ import annotations

# 线下"绩效设定表"Excel 解析
# 模板结构（sheet "目标设定表"，大量合并单元格）：
#   第 2 行：A2="姓名"（B2 起为值）、E2="工号"（F2 起为值）
#   第 6 行："第一部分：业绩目标"；第 7-8 行表头
#   第 9-13 行：目标行（A=序号 1-5，B=考核项，C=权重，D=关键成果及交付）
#   第 14 行：合计行（含公式，跳过）
import io
from dataclasses import dataclass, field

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlmodel import Session, select

from pms.database.models.user import User

SHEET_NAME = "目标设定表"

# 表头固定行
_NAME_ROW = 2        # A2=姓名 / E2=工号
_NAME_LABEL_COL = 1  # A 列
_EMP_ID_LABEL_COL = 5  # E 列
_OBJECTIVE_FIRST_ROW = 9
_OBJECTIVE_LAST_ROW = 13  # 第 14 行为合计行，不解析


@dataclass
class ParsedObjective:
    title: str
    weight: int  # 归一化为百分制整数（0.4 → 40）
    measure_criteria: str


@dataclass
class ParsedObjectiveSheet:
    wecom_userid: str
    name: str
    objectives: list[ParsedObjective] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def _merged_cell_value(ws: Worksheet, row: int, col: int) -> object:
    """读取单元格值；若落在合并区域内，取区域左上角的值。"""
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value
    coord = cell.coordinate
    for merged in ws.merged_cells.ranges:
        if coord in merged:
            return ws.cell(row=merged.min_row, column=merged.min_col).value
    return None


def _first_value_after_label(ws: Worksheet, row: int, label_col: int) -> str:
    """取标签列之后该行第一个非空单元格的值（值可能在合并区域内）。"""
    for col in range(label_col + 1, ws.max_column + 1):
        value = _merged_cell_value(ws, row, col)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def _normalize_weight(raw: object) -> int | None:
    """权重归一化：<=1 视为小数比例（0.4→40），>1 视为百分制（40→40）。"""
    if raw is None:
        return None
    try:
        value = float(str(raw).strip().rstrip("%"))
    except (ValueError, TypeError):
        return None
    if value <= 0:
        return None
    if value <= 1:
        return round(value * 100)
    return round(value)


def parse_offline_objective_sheet(file_bytes: bytes) -> ParsedObjectiveSheet:
    """解析线下目标设定表，返回工号/姓名/目标列表。

    解析失败（非 xlsx、缺 sheet 等）抛 ValueError，由调用方转为 skip。
    """
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise ValueError(f"文件格式错误：{e}") from e

    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active

    name = _first_value_after_label(ws, _NAME_ROW, _NAME_LABEL_COL)
    wecom_userid = _first_value_after_label(ws, _NAME_ROW, _EMP_ID_LABEL_COL)

    result = ParsedObjectiveSheet(wecom_userid=wecom_userid, name=name)

    for row in range(_OBJECTIVE_FIRST_ROW, _OBJECTIVE_LAST_ROW + 1):
        title_raw = _merged_cell_value(ws, row, 2)  # B=考核项
        if title_raw is None or not str(title_raw).strip():
            continue  # 空行跳过
        title = str(title_raw).strip()
        measure_raw = _merged_cell_value(ws, row, 4)  # D=关键成果及交付
        measure = str(measure_raw).strip() if measure_raw is not None else ""
        weight = _normalize_weight(_merged_cell_value(ws, row, 3))  # C=权重
        if weight is None:
            result.warnings.append(f"第{row}行目标「{title}」权重缺失或非法，按 0 处理")
            weight = 0
        result.objectives.append(
            ParsedObjective(title=title, weight=weight, measure_criteria=measure)
        )

    # 权重合计校验：!=100 只警告不拒绝（线下数据宽容）
    if result.objectives:
        total = sum(o.weight for o in result.objectives)
        if total != 100:
            result.warnings.append(f"权重合计为 {total}%，应为 100%")

    return result


def match_user_by_id_or_name(
    session: Session, wecom_userid: str, name: str
) -> tuple[User | None, str | None]:
    """线下表格用户匹配：工号非空按工号，否则按姓名匹配 active 用户。

    返回 (user, None) 表示匹配成功；(None, reason) 表示失败，reason 为 skip 原因。
    excel_import / probation_import 两个线下导入接口共用，避免逻辑分叉。
    """
    if wecom_userid:
        user = session.exec(
            select(User).where(User.wecom_userid == wecom_userid)
        ).first()
        if not user:
            return None, f"员工ID {wecom_userid} 不存在"
        return user, None
    if not name:
        return None, "未解析到工号和姓名，无法匹配员工"
    candidates = session.exec(
        select(User).where(User.name == name, User.status == "active")
    ).all()
    if len(candidates) == 1:
        return candidates[0], None
    if not candidates:
        return None, "姓名未匹配"
    return None, "姓名重名，请在表格中补充工号"
