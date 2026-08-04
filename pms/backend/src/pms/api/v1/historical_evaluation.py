from __future__ import annotations

# 历史绩效敏感数据：Excel 导入 + 严格权限查询
# 权限红线：仅 HR（hrbp/super_admin，按 visible_user_ids scope）与
# 员工直属上级（direct_leader 且 leader_userid 匹配）可见；
# 员工本人、dept_leader（含非直属）及其他角色一律 403
import io
import json
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel
from sqlmodel import Session, select

from pms.database.models.historical_evaluation import (
    HistoricalEvaluationDetail,
    HistoricalEvaluationSummary,
)
from pms.database.models.user import User
from pms.database.session import get_session
from pms.services.auth import get_current_user, require_role
from pms.services.scope import visible_user_ids
from pms.utils.audit import write_audit

import_router = APIRouter(prefix="/import/historical-evaluations", tags=["import"])
query_router = APIRouter(prefix="/history", tags=["history"])

SUMMARY_TEMPLATE_HEADERS = [
    "姓名",
    "工号",
    "周期名称",
    "上级绩效得分",
    "上级绩效等级",
    "上级价值观等级",
    "互评绩效得分",
    "互评绩效等级",
    "互评价值观等级",
    "自评绩效得分",
    "自评绩效等级",
    "自评价值观等级",
    "是否校准",
    "校准建议",
    "校准后绩效得分",
    "校准后结果",
    "备注",
]

DETAIL_TEMPLATE_HEADERS = [
    "姓名",
    "工号",
    "周期名称",
    "自评绩效得分",
    "自评价值观得分",
    "自评产出",
    "自评整体评价",
    "上级绩效得分",
    "上级价值观得分",
    "上级评价（汇总）",
    "互评1得分",
    "互评1评语",
    "互评2得分",
    "互评2评语",
    "互评3得分",
    "互评3评语",
    "互评4得分",
    "互评4评语",
    "互评5得分",
    "互评5评语",
]

MAX_PEERS = 5


class SkipItem(BaseModel):
    wecom_userid: str
    name: str
    reason: str


class ImportResult(BaseModel):
    imported: int
    skipped: list[SkipItem]


def _to_str(value: Any) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _to_float(value: Any) -> float | None:
    s = _to_str(value)
    if s is None:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _to_bool(value: Any) -> bool:
    s = _to_str(value)
    return s is not None and s.lower() in ("是", "true", "1", "yes", "y")


def _col(row: tuple, idx: int) -> Any:
    return row[idx] if len(row) > idx else None


def _load_rows(content: bytes) -> list[tuple]:
    try:
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"文件格式错误：{e}") from e
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    # 过滤整行为空的行
    rows = [r for r in rows if any(c is not None and str(c).strip() for c in r)]
    if not rows:
        raise HTTPException(status_code=400, detail="文件中无数据行")
    return rows


def _download_template(headers: list[str], sample: list[str], sheet: str, filename: str) -> StreamingResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(headers)
    ws.append(sample)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@import_router.get("/summary/template")
def download_summary_template() -> StreamingResponse:
    return _download_template(
        SUMMARY_TEMPLATE_HEADERS,
        ["张 Alice", "mock-alice", "2024H1", "3.75", "meet", "yi", "3.50", "meet", "yi",
         "3.50", "meet", "yi", "否", "", "", "", "表现稳定"],
        "历史绩效汇总导入",
        "historical_evaluation_summary_template.xlsx",
    )


@import_router.get("/detail/template")
def download_detail_template() -> StreamingResponse:
    return _download_template(
        DETAIL_TEMPLATE_HEADERS,
        ["张 Alice", "mock-alice", "2024H1", "3.50", "yi", "完成核心模块交付", "整体符合预期",
         "3.75", "yi", "表现稳定", "3.50", "协作顺畅", "4.00", "主动补位"],
        "历史绩效明细导入",
        "historical_evaluation_detail_template.xlsx",
    )


@import_router.post("/summary", response_model=ImportResult)
def import_summary(
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
    hr: User = Depends(require_role("hrbp", "super_admin")),
) -> ImportResult:
    rows = _load_rows(file.file.read())

    imported = 0
    skipped: list[SkipItem] = []

    for row in rows:
        name = _to_str(_col(row, 0)) or ""
        userid = _to_str(_col(row, 1)) or ""
        cycle_name = _to_str(_col(row, 2))

        if not userid:
            skipped.append(SkipItem(wecom_userid="", name=name, reason="工号为空"))
            continue
        if not cycle_name:
            skipped.append(SkipItem(wecom_userid=userid, name=name, reason="周期名称为空"))
            continue

        user = session.exec(select(User).where(User.wecom_userid == userid)).first()
        if not user:
            skipped.append(SkipItem(wecom_userid=userid, name=name, reason="工号不存在"))
            continue

        # 幂等：同 user + cycle 先删后插
        existing = session.exec(
            select(HistoricalEvaluationSummary).where(
                HistoricalEvaluationSummary.user_id == user.id,
                HistoricalEvaluationSummary.cycle_name == cycle_name,
            )
        ).first()
        if existing:
            session.delete(existing)
            session.flush()  # 先落删除，避免与后续插入撞唯一约束

        session.add(HistoricalEvaluationSummary(
            user_id=user.id,
            cycle_name=cycle_name,
            superior_score=_to_float(_col(row, 3)),
            superior_level=_to_str(_col(row, 4)),
            superior_value_grade=_to_str(_col(row, 5)),
            peer_avg_score=_to_float(_col(row, 6)),
            peer_level=_to_str(_col(row, 7)),
            peer_value_grade=_to_str(_col(row, 8)),
            self_score=_to_float(_col(row, 9)),
            self_level=_to_str(_col(row, 10)),
            self_value_grade=_to_str(_col(row, 11)),
            is_calibrated=_to_bool(_col(row, 12)),
            calibration_suggestion=_to_str(_col(row, 13)),
            calibrated_score=_to_float(_col(row, 14)),
            calibrated_result=_to_str(_col(row, 15)),
            comment=_to_str(_col(row, 16)),
        ))
        imported += 1

    if imported:
        write_audit(
            session,
            operator_userid=hr.wecom_userid,
            operator_name=hr.name,
            action="import_historical_evaluation_summary",
            resource_type="historical_evaluation_summary",
            resource_id="",
            after={"imported": imported, "skipped": len(skipped)},
        )
    session.commit()

    return ImportResult(imported=imported, skipped=skipped)


@import_router.post("/detail", response_model=ImportResult)
def import_detail(
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
    hr: User = Depends(require_role("hrbp", "super_admin")),
) -> ImportResult:
    rows = _load_rows(file.file.read())

    imported = 0
    skipped: list[SkipItem] = []

    for row in rows:
        name = _to_str(_col(row, 0)) or ""
        userid = _to_str(_col(row, 1)) or ""
        cycle_name = _to_str(_col(row, 2))

        if not userid:
            skipped.append(SkipItem(wecom_userid="", name=name, reason="工号为空"))
            continue
        if not cycle_name:
            skipped.append(SkipItem(wecom_userid=userid, name=name, reason="周期名称为空"))
            continue

        user = session.exec(select(User).where(User.wecom_userid == userid)).first()
        if not user:
            skipped.append(SkipItem(wecom_userid=userid, name=name, reason="工号不存在"))
            continue

        # 互评明细：最多 5 条，只存非空（得分或评语任一非空）条目
        peers: list[dict[str, Any]] = []
        for i in range(MAX_PEERS):
            score = _to_float(_col(row, 10 + i * 2))
            comment = _to_str(_col(row, 11 + i * 2))
            if score is None and comment is None:
                continue
            peers.append({"index": i + 1, "score": score, "comment": comment})

        # 幂等：同 user + cycle 先删后插
        existing = session.exec(
            select(HistoricalEvaluationDetail).where(
                HistoricalEvaluationDetail.user_id == user.id,
                HistoricalEvaluationDetail.cycle_name == cycle_name,
            )
        ).first()
        if existing:
            session.delete(existing)
            session.flush()  # 先落删除，避免与后续插入撞唯一约束

        session.add(HistoricalEvaluationDetail(
            user_id=user.id,
            cycle_name=cycle_name,
            self_score=_to_float(_col(row, 3)),
            self_value_grade=_to_str(_col(row, 4)),
            self_output=_to_str(_col(row, 5)),
            self_comment=_to_str(_col(row, 6)),
            superior_score=_to_float(_col(row, 7)),
            superior_value_grade=_to_str(_col(row, 8)),
            superior_comment=_to_str(_col(row, 9)),
            peers_json=json.dumps(peers, ensure_ascii=False) if peers else None,
        ))
        imported += 1

    if imported:
        write_audit(
            session,
            operator_userid=hr.wecom_userid,
            operator_name=hr.name,
            action="import_historical_evaluation_detail",
            resource_type="historical_evaluation_detail",
            resource_id="",
            after={"imported": imported, "skipped": len(skipped)},
        )
    session.commit()

    return ImportResult(imported=imported, skipped=skipped)


def _summary_view(s: HistoricalEvaluationSummary) -> dict[str, Any]:
    return {
        "superior_score": s.superior_score,
        "superior_level": s.superior_level,
        "superior_value_grade": s.superior_value_grade,
        "peer_avg_score": s.peer_avg_score,
        "peer_level": s.peer_level,
        "peer_value_grade": s.peer_value_grade,
        "self_score": s.self_score,
        "self_level": s.self_level,
        "self_value_grade": s.self_value_grade,
        "is_calibrated": s.is_calibrated,
        "calibration_suggestion": s.calibration_suggestion,
        "calibrated_score": s.calibrated_score,
        "calibrated_result": s.calibrated_result,
        "comment": s.comment,
    }


def _detail_view(d: HistoricalEvaluationDetail) -> dict[str, Any]:
    peers: list[dict[str, Any]] = []
    if d.peers_json:
        try:
            parsed = json.loads(d.peers_json)
            if isinstance(parsed, list):
                peers = parsed
        except json.JSONDecodeError:
            peers = []
    return {
        "self_score": d.self_score,
        "self_value_grade": d.self_value_grade,
        "self_output": d.self_output,
        "self_comment": d.self_comment,
        "superior_score": d.superior_score,
        "superior_value_grade": d.superior_value_grade,
        "superior_comment": d.superior_comment,
        "peers": peers,
    }


@query_router.get("/users/{user_id}/evaluations")
def list_user_historical_evaluations(
    user_id: int,
    session: Session = Depends(get_session),
    current: User = Depends(get_current_user),
) -> list[dict[str, Any]]:
    """查看某员工全部历史周期的绩效汇总+明细（敏感数据，严格权限）。

    仅 HR（hrbp/super_admin，按数据可见范围）与直属上级
    （direct_leader 且 leader_userid 匹配）可见；其余一律 403。
    """
    target = session.get(User, user_id)
    if not target:
        raise HTTPException(status_code=404, detail="用户不存在")

    if current.role in ("hrbp", "super_admin"):
        scope = visible_user_ids(session, current)
        if scope is not None and target.id not in scope:
            raise HTTPException(status_code=403, detail="该用户不在你的数据可见范围内")
    elif current.role == "direct_leader":
        if target.leader_userid != current.wecom_userid:
            raise HTTPException(status_code=403, detail="仅直属上级可查看历史绩效敏感数据")
    else:
        raise HTTPException(status_code=403, detail="仅 HR 或直属上级可查看历史绩效敏感数据")

    summaries = session.exec(
        select(HistoricalEvaluationSummary).where(HistoricalEvaluationSummary.user_id == user_id)
    ).all()
    details = session.exec(
        select(HistoricalEvaluationDetail).where(HistoricalEvaluationDetail.user_id == user_id)
    ).all()

    summary_map = {s.cycle_name: s for s in summaries}
    detail_map = {d.cycle_name: d for d in details}
    cycle_names = sorted(set(summary_map) | set(detail_map))

    return [
        {
            "cycle_name": name,
            "summary": _summary_view(summary_map[name]) if name in summary_map else None,
            "detail": _detail_view(detail_map[name]) if name in detail_map else None,
        }
        for name in cycle_names
    ]
