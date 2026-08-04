from __future__ import annotations

# 试用期目标批量导入（HRBP 线下收集 → Excel 批量覆盖写入）
# 参考 excel_import.py / historical_import.py 的解析与模板模式
import io
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlmodel import Session, select

from pms.database.models.enums import ProbationObjectiveStatus, ProbationPlanStatus
from pms.database.models.probation import ProbationObjective, ProbationPlan
from pms.database.models.user import User
from pms.database.session import get_session
from pms.services.auth import require_role
from pms.services.offline_template import match_user_by_id_or_name, parse_offline_objective_sheet
from pms.utils.audit import write_audit

router = APIRouter(prefix="/probation", tags=["probation"])

TEMPLATE_HEADERS = [
    "员工ID（wecom_userid）",
    "姓名（校对用）",
    "目标项",
    "目标描述",
    "衡量标准",
]

# 允许覆盖导入的计划状态（与 save_probation_objectives 一致）
_IMPORTABLE_STATUSES = (
    ProbationPlanStatus.DRAFT,
    ProbationPlanStatus.OBJECTIVE_DRAFT,
    ProbationPlanStatus.OBJECTIVE_PENDING_REVIEW,
)


@router.get("/import-objectives/template")
def download_import_template():
    # 生成空白 Excel 模板供 HRBP 下载（与现有模板接口一致，无需鉴权）
    wb = Workbook()
    ws = wb.active
    ws.title = "试用期目标导入"
    ws.append(TEMPLATE_HEADERS)
    ws.append([
        "mock-alice", "张 Alice",
        "完成 V1.0 MVP 上线", "按期交付所有功能模块", "12月底前全员使用",
    ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=probation_objectives_template.xlsx"},
    )


@router.post("/import-objectives")
def import_probation_objectives(
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
    hr: User = Depends(require_role("hrbp", "super_admin")),
):
    try:
        content = file.file.read()
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"文件格式错误：{e}") from e

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="文件中无数据行")

    # 按员工分组（保持 Excel 行序）
    grouped: dict[str, dict] = {}
    for row in rows:
        uid = str(row[0]).strip() if len(row) > 0 and row[0] else ""
        name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        title = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        desc = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        measure = str(row[4]).strip() if len(row) > 4 and row[4] else ""
        if not uid:
            continue
        entry = grouped.setdefault(uid, {"name": name, "objectives": []})
        entry["objectives"].append({
            "title": title,
            "description": desc,
            "measure_criteria": measure,
        })

    skipped: list[dict] = []
    to_import: list[dict] = []  # {"user", "plan", "objectives"}

    for uid, entry in grouped.items():
        name = entry["name"]
        objectives = entry["objectives"]

        skip_reason: str | None = None
        user = session.exec(select(User).where(User.wecom_userid == uid)).first()
        if not user:
            skip_reason = "员工ID 不存在"
        else:
            plan = session.exec(
                select(ProbationPlan).where(ProbationPlan.user_id == user.id)
            ).first()
            if not plan:
                skip_reason = "试用期计划不存在"
            elif plan.status not in _IMPORTABLE_STATUSES:
                skip_reason = f"计划状态 {plan.status} 不允许导入目标"
            elif not 1 <= len(objectives) <= 10:
                # 校验与 save_probation_objectives 一致：1-10 条
                skip_reason = f"目标数量 {len(objectives)} 不在 1-10 范围内"
            else:
                invalid = next(
                    (i + 1 for i, o in enumerate(objectives)
                     if not o["title"] or not o["description"] or not o["measure_criteria"]),
                    None,
                )
                if invalid is not None:
                    skip_reason = f"第 {invalid} 条目标的标题/描述/衡量标准不能为空"

        if skip_reason:
            skipped.append({"wecom_userid": uid, "name": name, "reason": skip_reason})
            continue

        to_import.append({"user": user, "plan": plan, "objectives": objectives})

    # 覆盖写入：先删旧目标再插入（一个事务）
    now = datetime.now(timezone.utc)
    imported_objectives = 0
    for item in to_import:
        plan = item["plan"]
        old = session.exec(
            select(ProbationObjective).where(ProbationObjective.plan_id == plan.id)
        ).all()
        for o in old:
            session.delete(o)
        session.flush()

        for i, o in enumerate(item["objectives"]):
            session.add(ProbationObjective(
                plan_id=plan.id,
                title=o["title"],
                description=o["description"],
                measure_criteria=o["measure_criteria"],
                order_num=i,
                status=ProbationObjectiveStatus.DRAFT,
            ))
        imported_objectives += len(item["objectives"])

        if plan.status == ProbationPlanStatus.DRAFT:
            plan.status = ProbationPlanStatus.OBJECTIVE_DRAFT
        plan.updated_at = now
        session.add(plan)

    if to_import:
        write_audit(
            session,
            operator_userid=hr.wecom_userid,
            operator_name=hr.name,
            action="import_probation_objectives",
            resource_type="probation_plan",
            resource_id="-",
            after={
                "imported_users": len(to_import),
                "imported_objectives": imported_objectives,
                "skipped_count": len(skipped),
            },
        )
    session.commit()

    return {
        "imported_users": len(to_import),
        "imported_objectives": imported_objectives,
        "skipped": skipped,
    }


@router.post("/import-offline-objectives")
def import_offline_probation_objectives(
    files: list[UploadFile] = File(...),
    session: Session = Depends(get_session),
    hr: User = Depends(require_role("hrbp", "super_admin")),
):
    # 线下"绩效设定表"多文件导入：每文件 = 一名员工，覆盖写入试用期目标
    skipped: list[dict] = []
    warnings: list[str] = []
    to_import: list[dict] = []  # {"user", "plan", "objectives"}

    for f in files:
        label = f.filename or "未命名文件"
        try:
            sheet = parse_offline_objective_sheet(f.file.read())
        except ValueError as e:
            skipped.append({"wecom_userid": "", "name": label, "reason": str(e)})
            continue
        warnings.extend(f"{label}: {w}" for w in sheet.warnings)

        skip_reason: str | None = None
        user = None
        plan = None
        if not sheet.objectives:
            skip_reason = "未解析到目标"
        else:
            # 工号非空按工号，否则按姓名匹配 active 用户（共用 match_user_by_id_or_name）
            user, skip_reason = match_user_by_id_or_name(
                session, sheet.wecom_userid, sheet.name
            )
            if user:
                plan = session.exec(
                    select(ProbationPlan).where(ProbationPlan.user_id == user.id)
                ).first()
                if not plan:
                    skip_reason = "试用期计划不存在"
                elif plan.status not in _IMPORTABLE_STATUSES:
                    skip_reason = f"计划状态 {plan.status} 不允许导入目标"

        if skip_reason:
            skipped.append({
                "wecom_userid": sheet.wecom_userid,
                "name": sheet.name or label,
                "reason": skip_reason,
            })
            continue
        to_import.append({"user": user, "plan": plan, "objectives": sheet.objectives})

    # 覆盖写入：先删旧目标再插入（一个事务）
    now = datetime.now(timezone.utc)
    imported_objectives = 0
    for item in to_import:
        plan = item["plan"]
        old = session.exec(
            select(ProbationObjective).where(ProbationObjective.plan_id == plan.id)
        ).all()
        for o in old:
            session.delete(o)
        session.flush()

        for i, o in enumerate(item["objectives"]):
            session.add(ProbationObjective(
                plan_id=plan.id,
                title=o.title,
                description=o.measure_criteria,
                measure_criteria=o.measure_criteria,
                weight=o.weight,
                order_num=i,
                status=ProbationObjectiveStatus.DRAFT,
            ))
        imported_objectives += len(item["objectives"])

        if plan.status == ProbationPlanStatus.DRAFT:
            plan.status = ProbationPlanStatus.OBJECTIVE_DRAFT
        plan.updated_at = now
        session.add(plan)

    if to_import:
        write_audit(
            session,
            operator_userid=hr.wecom_userid,
            operator_name=hr.name,
            action="import_offline_probation_objectives",
            resource_type="probation_plan",
            resource_id="-",
            after={
                "imported_users": len(to_import),
                "imported_objectives": imported_objectives,
                "skipped_count": len(skipped),
            },
        )
    session.commit()

    return {
        "imported_users": len(to_import),
        "imported_objectives": imported_objectives,
        "skipped": skipped,
        "warnings": warnings,
    }
