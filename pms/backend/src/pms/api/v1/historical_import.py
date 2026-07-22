from __future__ import annotations

# 历史绩效结果批量导入（只读快照，不参与当前流程）
import io
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel
from sqlmodel import Session, select

from pms.database.models.historical_performance import HistoricalPerformanceResult
from pms.database.models.user import User
from pms.database.session import get_session
from pms.services.auth import require_role
from pms.utils.audit import write_audit

router = APIRouter(prefix="/import/historical-performance", tags=["import"])

TEMPLATE_HEADERS = [
    "员工ID（wecom_userid）",
    "姓名（校对用）",
    "周期名称",
    "业绩分（1-5，0.25分段）",
    "业绩等级",
    "价值观-信念",
    "价值观-团队",
    "价值观-成长",
    "上级评语",
]


@router.get("/template")
def download_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "历史绩效导入"
    ws.append(TEMPLATE_HEADERS)
    ws.append([
        "mock-alice", "张 Alice", "2024H1", "3.75", "meet", "yi", "jia", "yi", "表现稳定",
    ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=historical_performance_template.xlsx"},
    )


class ImportResult(BaseModel):
    total: int
    success: int
    failed: int
    errors: list[str]


@router.post("", response_model=ImportResult)
def import_historical_performance(
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
    hr: User = Depends(require_role("hrbp", "super_admin")),
):
    try:
        content = file.file.read()
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"文件格式错误：{e}") from e

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="文件中无数据行")

    errors: list[str] = []
    success = 0
    failed = 0

    for i, row in enumerate(rows, start=2):
        if len(row) < 8:
            errors.append(f"第{i}行：列数不足")
            failed += 1
            continue

        uid = str(row[0]).strip() if row[0] else ""
        cycle_name = str(row[2]).strip() if row[2] else ""
        perf_score_raw = row[3]
        perf_level = str(row[4]).strip() if row[4] else None
        value_belief = str(row[5]).strip() if row[5] else None
        value_team = str(row[6]).strip() if row[6] else None
        value_growth = str(row[7]).strip() if row[7] else None
        comment = str(row[8]).strip() if len(row) > 8 and row[8] else None

        if not uid:
            errors.append(f"第{i}行：员工ID 为空")
            failed += 1
            continue
        if not cycle_name:
            errors.append(f"第{i}行：周期名称 为空")
            failed += 1
            continue

        # 校验业绩分
        perf_score: float | None = None
        if perf_score_raw is not None and perf_score_raw != "":
            try:
                perf_score = float(str(perf_score_raw))
                if perf_score < 1 or perf_score > 5:
                    raise ValueError
                # 0.25 分段校验
                if abs(perf_score * 4 - round(perf_score * 4)) > 1e-9:
                    raise ValueError
            except (ValueError, TypeError):
                errors.append(f"第{i}行：业绩分必须为 1-5 之间且 0.25 分段")
                failed += 1
                continue

        # 校验员工存在
        user = session.exec(select(User).where(User.wecom_userid == uid)).first()
        if not user:
            errors.append(f"第{i}行：员工ID {uid} 不存在")
            failed += 1
            continue

        # 检查是否重复导入
        existing = session.exec(
            select(HistoricalPerformanceResult).where(
                HistoricalPerformanceResult.user_id == user.id,
                HistoricalPerformanceResult.cycle_name == cycle_name,
            )
        ).first()
        if existing:
            errors.append(f"第{i}行：{uid} 的 {cycle_name} 已存在，跳过")
            failed += 1
            continue

        record = HistoricalPerformanceResult(
            user_id=user.id,
            cycle_name=cycle_name,
            perf_score=perf_score,
            perf_level=perf_level,
            value_belief=value_belief,
            value_team=value_team,
            value_growth=value_growth,
            comment=comment,
            imported_by=hr.wecom_userid,
        )
        session.add(record)
        success += 1

    if success:
        write_audit(
            session,
            operator_userid=hr.wecom_userid,
            operator_name=hr.name,
            action="import_historical_performance",
            resource_type="historical_performance",
            resource_id="",
            after={"success": success, "failed": failed},
        )
    session.commit()

    return ImportResult(total=len(rows), success=success, failed=failed, errors=errors)


@router.get("", response_model=list[dict])
def list_historical_performance(
    user_id: int | None = None,
    session: Session = Depends(get_session),
    hr: User = Depends(require_role("hrbp", "super_admin")),
):
    q = select(HistoricalPerformanceResult, User).join(User, User.id == HistoricalPerformanceResult.user_id)
    if user_id:
        q = q.where(HistoricalPerformanceResult.user_id == user_id)
    rows = session.exec(q.order_by(HistoricalPerformanceResult.created_at.desc())).all()
    return [
        {
            "id": r.id,
            "user_id": r.user_id,
            "user_name": u.name,
            "cycle_name": r.cycle_name,
            "perf_score": r.perf_score,
            "perf_level": r.perf_level,
            "value_belief": r.value_belief,
            "value_team": r.value_team,
            "value_growth": r.value_growth,
            "comment": r.comment,
            "imported_by": r.imported_by,
            "created_at": r.created_at.isoformat(),
        }
        for r, u in rows
    ]
