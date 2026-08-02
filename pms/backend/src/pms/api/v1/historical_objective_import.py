from __future__ import annotations

# 历史（线下）绩效目标留档批量导入（只读快照，不参与当前流程）
# 幂等策略：同一员工 + 周期名称先删后插
import io

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlmodel import Session, select

from pms.database.models.historical_objective import HistoricalObjective
from pms.database.models.user import User
from pms.database.session import get_session
from pms.services.auth import get_current_user, require_role
from pms.services.scope import visible_user_ids
from pms.utils.audit import write_audit

router = APIRouter(prefix="/import/historical-objectives", tags=["import"])

TEMPLATE_HEADERS = [
    "员工ID（wecom_userid）",
    "姓名（校对用）",
    "周期名称",
    "目标项",
    "目标描述",
    "衡量标准",
    "权重(%)",
]


@router.get("/template")
def download_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "历史目标导入"
    ws.append(TEMPLATE_HEADERS)
    ws.append([
        "mock-alice", "张 Alice", "2024H1",
        "完成 V1.0 MVP 上线", "按期交付所有功能模块", "12月底前全员使用", "40",
    ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=historical_objectives_template.xlsx"},
    )


@router.post("")
def import_historical_objectives(
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
    hr: User = Depends(require_role("hrbp", "super_admin")),
):
    try:
        content = file.file.read()
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"文件格式错误：{e}") from e

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="文件中无数据行")

    # 按 (员工ID, 周期名称) 分组（保持 Excel 行序）
    grouped: dict[tuple[str, str], dict] = {}
    for row in rows:
        uid = str(row[0]).strip() if len(row) > 0 and row[0] else ""
        name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        cycle_name = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        title = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        desc = str(row[4]).strip() if len(row) > 4 and row[4] else ""
        measure = str(row[5]).strip() if len(row) > 5 and row[5] else ""
        weight_raw = row[6] if len(row) > 6 else None

        weight = 0
        if weight_raw is not None and str(weight_raw).strip() != "":
            try:
                weight = int(float(str(weight_raw)))
            except (ValueError, TypeError):
                weight = -1  # 标记非法，分组后统一处理

        if not uid:
            continue
        key = (uid, cycle_name)
        entry = grouped.setdefault(key, {"name": name, "objectives": []})
        entry["objectives"].append({
            "title": title,
            "description": desc,
            "measure_criteria": measure,
            "weight": weight,
        })

    skipped: list[dict] = []
    to_import: list[dict] = []  # {"user", "cycle_name", "objectives"}

    for (uid, cycle_name), entry in grouped.items():
        name = entry["name"]
        objectives = entry["objectives"]

        skip_reason: str | None = None
        if not cycle_name:
            skip_reason = "周期名称 为空"
        else:
            user = session.exec(select(User).where(User.wecom_userid == uid)).first()
            if not user:
                skip_reason = "员工ID 不存在"
            else:
                invalid = next(
                    (i + 1 for i, o in enumerate(objectives) if not o["title"]),
                    None,
                )
                if invalid is not None:
                    skip_reason = f"第 {invalid} 条目标的目标项不能为空"
                elif any(o["weight"] < 0 for o in objectives):
                    skip_reason = "权重必须为整数"

        if skip_reason:
            skipped.append({"wecom_userid": uid, "name": name, "reason": skip_reason})
            continue

        to_import.append({"user": user, "cycle_name": cycle_name, "objectives": objectives})

    # 幂等覆盖：同一员工 + 周期名称先删后插（整批一个事务）
    imported_objectives = 0
    imported_user_ids: set[int] = set()
    for item in to_import:
        user = item["user"]
        cycle_name = item["cycle_name"]
        old = session.exec(
            select(HistoricalObjective).where(
                HistoricalObjective.user_id == user.id,
                HistoricalObjective.cycle_name == cycle_name,
            )
        ).all()
        for o in old:
            session.delete(o)
        session.flush()

        for i, o in enumerate(item["objectives"]):
            session.add(HistoricalObjective(
                user_id=user.id,
                cycle_name=cycle_name,
                title=o["title"],
                description=o["description"] or None,
                measure_criteria=o["measure_criteria"] or None,
                weight=o["weight"],
                order_num=i,
            ))
        imported_objectives += len(item["objectives"])
        imported_user_ids.add(user.id)

    if to_import:
        write_audit(
            session,
            operator_userid=hr.wecom_userid,
            operator_name=hr.name,
            action="import_historical_objectives",
            resource_type="historical_objective",
            resource_id="-",
            after={
                "imported_users": len(imported_user_ids),
                "imported_objectives": imported_objectives,
                "skipped_count": len(skipped),
            },
        )
    session.commit()

    return {
        "imported_users": len(imported_user_ids),
        "imported_objectives": imported_objectives,
        "skipped": skipped,
    }


@router.get("", response_model=list[dict])
def list_historical_objectives(
    user_id: int | None = None,
    session: Session = Depends(get_session),
    current: User = Depends(get_current_user),
):
    # 查询口径与 historical_import.py GET 一致：
    # hrbp/super_admin 按数据可见范围过滤，其余角色只看自己
    q = select(HistoricalObjective, User).join(User, User.id == HistoricalObjective.user_id)
    if current.role in ("hrbp", "super_admin"):
        scope = visible_user_ids(session, current)
        if scope is not None:
            q = q.where(HistoricalObjective.user_id.in_(scope))
    else:
        q = q.where(HistoricalObjective.user_id == current.id)
    if user_id:
        q = q.where(HistoricalObjective.user_id == user_id)
    rows = session.exec(
        q.order_by(HistoricalObjective.cycle_name, HistoricalObjective.order_num)
    ).all()
    return [
        {
            "id": r.id,
            "user_id": r.user_id,
            "user_name": u.name,
            "cycle_name": r.cycle_name,
            "title": r.title,
            "description": r.description,
            "measure_criteria": r.measure_criteria,
            "weight": r.weight,
            "order_num": r.order_num,
        }
        for r, u in rows
    ]
