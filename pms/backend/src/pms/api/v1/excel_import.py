from __future__ import annotations

# Excel 目标批量导入（PRD 3.3.1 一次性过渡功能）
# HRBP 下载模板 → 填写 → 上传 → 系统校验 → 导入
import io
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlmodel import Session, select

from pms.database.models.objective import Objective
from pms.database.models.objective_cycle import ObjectiveCycle
from pms.database.models.objective_cycle_participant import ObjectiveCycleParticipant
from pms.database.models.user import Department, User
from pms.database.session import get_session
from pms.services.auth import require_role
from pms.utils.audit import write_audit

router = APIRouter(prefix="/objective-cycles", tags=["objective-cycles"])

# 模板列定义（PRD 3.3.1 字段规范）
TEMPLATE_HEADERS = [
    "员工ID（wecom_userid）",
    "姓名（校对用）",
    "目标类别",
    "目标项",
    "目标描述",
    "衡量标准",
    "权重(%)",
    "目标周期",
]


@router.get("/excel/template")
def download_template():
    # 生成空白 Excel 模板供 HRBP 下载
    wb = Workbook()
    ws = wb.active
    ws.title = "绩效目标导入"
    ws.append(TEMPLATE_HEADERS)
    # 写一行示例
    ws.append([
        "mock-alice", "张 Alice", "业绩目标",
        "完成 V1.0 MVP 上线", "按期交付所有功能模块",
        "12月底前全员使用", "40", "2025H2",
    ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=objective_import_template.xlsx"},
    )


@router.post("/{objective_cycle_id}/excel/import")
def import_objectives(
    objective_cycle_id: int,
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
    hr: User = Depends(require_role("hrbp", "super_admin")),
):
    # 校验 + 导入
    cycle = session.get(ObjectiveCycle, objective_cycle_id)
    if not cycle:
        raise HTTPException(status_code=404, detail="目标周期不存在")
    if cycle.status not in ("draft", "active"):
        raise HTTPException(status_code=400, detail="当前目标周期状态不允许导入目标")

    # 读取上传的 Excel
    try:
        content = file.file.read()
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"文件格式错误：{e}") from e

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="文件中无数据行")

    # 校验每一行
    errors: list[str] = []
    parsed: list[dict] = []
    # 用于校验同一员工权重之和
    weight_map: dict[str, int] = {}

    for i, row in enumerate(rows, start=2):
        if len(row) < 7:
            errors.append(f"第{i}行：列数不足")
            continue
        uid, name, category, title, desc, measure, weight_raw, *_ = row
        uid = str(uid).strip() if uid else ""
        title = str(title).strip() if title else ""
        desc = str(desc).strip() if desc else ""
        measure = str(measure).strip() if measure else ""

        if not uid:
            errors.append(f"第{i}行：员工ID 为空")
            continue
        if not title:
            errors.append(f"第{i}行：目标项 为空")
            continue
        if not desc:
            errors.append(f"第{i}行：目标描述 为空")
            continue
        if not measure:
            errors.append(f"第{i}行：衡量标准 为空")
            continue

        # 权重
        try:
            weight = int(float(str(weight_raw)))
            if weight <= 0 or weight > 100:
                raise ValueError
        except (ValueError, TypeError):
            errors.append(f"第{i}行：权重必须为 1-100 的整数")
            continue

        # 员工存在性
        user = session.exec(select(User).where(User.wecom_userid == uid)).first()
        if not user:
            errors.append(f"第{i}行：员工ID {uid} 不存在")
            continue

        weight_map[uid] = weight_map.get(uid, 0) + weight
        parsed.append({
            "user_id": user.id,
            "uid": uid,
            "title": title,
            "description": desc,
            "measure_criteria": measure,
            "weight": weight,
            "category": str(category).strip() if category else None,
        })

    # 权重总和校验
    for uid, total in weight_map.items():
        if total != 100:
            errors.append(f"员工 {uid} 的权重总和为 {total}，应为 100")

    if errors:
        raise HTTPException(status_code=400, detail={"errors": errors})

    # 全部通过，写入数据库（先删旧目标再写新的）
    user_ids = list({p["user_id"] for p in parsed})
    existing = session.exec(
        select(Objective).where(
            Objective.objective_cycle_id == objective_cycle_id,
            Objective.user_id.in_(user_ids),
        )
    ).all()

    # 目标已被上级确认（approved/locked）的员工跳过导入，避免无差别覆盖
    locked_user_ids = {obj.user_id for obj in existing if obj.status in ("approved", "locked")}
    skipped_users = sorted({p["uid"] for p in parsed if p["user_id"] in locked_user_ids})
    if locked_user_ids:
        parsed = [p for p in parsed if p["user_id"] not in locked_user_ids]
        existing = [obj for obj in existing if obj.user_id not in locked_user_ids]
        user_ids = list({p["user_id"] for p in parsed})

    for obj in existing:
        session.delete(obj)
    session.flush()

    # 为导入的员工自动创建/更新参与人记录
    existing_participants = session.exec(
        select(ObjectiveCycleParticipant).where(
            ObjectiveCycleParticipant.objective_cycle_id == objective_cycle_id,
            ObjectiveCycleParticipant.user_id.in_(user_ids),
        )
    ).all()
    participant_user_ids = {p.user_id for p in existing_participants}
    for uid in user_ids:
        if uid in participant_user_ids:
            continue
        user = session.get(User, uid)
        if not user:
            continue
        dept = session.get(Department, user.department_id) if user.department_id else None
        session.add(ObjectiveCycleParticipant(
            objective_cycle_id=objective_cycle_id,
            user_id=uid,
            leader_userid_snapshot=user.leader_userid,
            dept_name_snapshot=dept.name if dept else None,
            status="approved",
        ))

    now = datetime.now(timezone.utc)
    for idx, p in enumerate(parsed):
        session.add(Objective(
            objective_cycle_id=objective_cycle_id,
            user_id=p["user_id"],
            title=p["title"],
            description=p["description"],
            measure_criteria=p["measure_criteria"],
            weight=p["weight"],
            order_num=idx,
            status="approved",          # Excel 导入视为已确认目标
            reviewed_by=hr.wecom_userid,
            reviewed_at=now,
        ))

    write_audit(
        session,
        operator_userid=hr.wecom_userid,
        operator_name=hr.name,
        action="excel_import_objectives",
        resource_type="objective",
        resource_id=str(objective_cycle_id),
        after={"row_count": len(parsed), "user_count": len(user_ids)},
    )
    session.commit()

    return {
        "imported_rows": len(parsed),
        "affected_users": len(user_ids),
        "skipped_users": skipped_users,
    }
