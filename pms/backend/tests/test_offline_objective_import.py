"""线下"绩效设定表"导入测试。

覆盖：
- 解析器：真实样本（填充后）工号/姓名/目标/权重归一化、权重合计 warning
- 目标周期导入：写入 draft、非参与人自动加入、approved 员工 skip
- 试用期导入：正常覆盖、无计划 skip、非 HR 403
"""

from __future__ import annotations

import io
from datetime import date, timedelta
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlmodel import Session, select

from pms.database.models.enums import ProbationPlanStatus
from pms.database.models.objective import Objective
from pms.database.models.objective_cycle_participant import ObjectiveCycleParticipant
from pms.database.models.probation import ProbationObjective, ProbationPlan
from pms.database.models.user import User
from pms.database.session import engine
from pms.main import app
from pms.services.offline_template import parse_offline_objective_sheet

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
FIXTURE = Path(__file__).parent / "fixtures" / "offline_objective_template.xlsx"


@pytest.fixture
def client():
    return TestClient(app)


def _login(client: TestClient, wecom_userid: str) -> str:
    resp = client.post("/api/v1/auth/mock-login", json={"wecom_userid": wecom_userid})
    assert resp.status_code == 200, resp.text
    return resp.json()["token"]


def _headers(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


def _make_offline_sheet(
    wecom_userid: str,
    name: str,
    objectives: list[tuple[str, object, str]],
) -> bytes:
    """在真实样本模板上填充 姓名/工号/目标行，返回 xlsx 字节。

    objectives: [(考核项, 权重, 关键成果)]，从第 9 行起依次写入。
    """
    wb = load_workbook(FIXTURE)
    ws = wb["目标设定表"]
    ws["B2"] = name  # B2:D2 合并区左上角
    ws["F2"] = wecom_userid  # F2:J2 合并区左上角
    for i, (title, weight, measure) in enumerate(objectives):
        row = 9 + i
        ws.cell(row=row, column=2, value=title)
        ws.cell(row=row, column=3, value=weight)
        ws.cell(row=row, column=4, value=measure)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload_cycle(
    client: TestClient, token: str, objective_cycle_id: int, files: list[tuple[str, bytes]]
):
    return client.post(
        f"/api/v1/objective-cycles/{objective_cycle_id}/excel/import-offline",
        headers=_headers(token),
        files=[("files", (fname, data, XLSX_MIME)) for fname, data in files],
    )


def _upload_probation(client: TestClient, token: str, files: list[tuple[str, bytes]]):
    return client.post(
        "/api/v1/probation/import-offline-objectives",
        headers=_headers(token),
        files=[("files", (fname, data, XLSX_MIME)) for fname, data in files],
    )


def _user_id(wecom_userid: str) -> int:
    with Session(engine) as s:
        user = s.exec(select(User).where(User.wecom_userid == wecom_userid)).first()
        assert user, f"seed 用户 {wecom_userid} 不存在"
        return user.id


def _create_objective_cycle(client: TestClient, hr_token: str, name: str) -> int:
    resp = client.post(
        "/api/v1/objective-cycles",
        headers=_headers(hr_token),
        json={"name": name, "start_date": "2025-07-01", "end_date": "2025-12-31"},
    )
    assert resp.status_code == 200, resp.text
    return resp.json()["id"]


def _reset_plan(wecom_userid: str, status: str = ProbationPlanStatus.DRAFT) -> int:
    """清空该用户的试用期计划并重建一个指定状态的计划，返回 plan.id。"""
    with Session(engine) as s:
        user = s.exec(select(User).where(User.wecom_userid == wecom_userid)).first()
        assert user, f"seed 用户 {wecom_userid} 不存在"
        old_plans = s.exec(select(ProbationPlan).where(ProbationPlan.user_id == user.id)).all()
        for p in old_plans:
            for o in s.exec(
                select(ProbationObjective).where(ProbationObjective.plan_id == p.id)
            ).all():
                s.delete(o)
            s.delete(p)
        # 同一 flush 内 SQLAlchemy 先 INSERT 后 DELETE，
        # 会撞 uq_probation_plan_user，先 flush 把删除落库再建新计划
        s.flush()
        plan = ProbationPlan(
            user_id=user.id,
            start_date=date.today() - timedelta(days=30),
            end_date=date.today() + timedelta(days=150),
            probation_months=6,
            status=status,
        )
        s.add(plan)
        s.commit()
        s.refresh(plan)
        return plan.id


# ---------- 解析器 ----------


def test_parse_sample_sheet() -> None:
    data = _make_offline_sheet(
        "mock-alice", "张 Alice",
        [("业绩KPI-交付", 0.4, "里程碑全部达成"),
         ("重点工作-协作", 0.3, "无重大投诉"),
         ("能力提升", 0.3, "落地案例 1 个")],
    )
    sheet = parse_offline_objective_sheet(data)
    assert sheet.wecom_userid == "mock-alice"
    assert sheet.name == "张 Alice"
    assert len(sheet.objectives) == 3
    assert [o.title for o in sheet.objectives] == ["业绩KPI-交付", "重点工作-协作", "能力提升"]
    # 小数权重归一化为百分制整数
    assert [o.weight for o in sheet.objectives] == [40, 30, 30]
    assert sheet.objectives[0].measure_criteria == "里程碑全部达成"
    assert sheet.warnings == []


def test_parse_percent_weight_form() -> None:
    # 百分制写法（>1）直接取整
    data = _make_offline_sheet("mock-bob", "李 Bob", [("目标一", 60, "标准一"), ("目标二", 40, "标准二")])
    sheet = parse_offline_objective_sheet(data)
    assert [o.weight for o in sheet.objectives] == [60, 40]
    assert sheet.warnings == []


def test_parse_weight_sum_warning() -> None:
    # 合计 != 100 只 warning 不拒绝
    data = _make_offline_sheet("mock-bob", "李 Bob", [("目标一", 0.4, "标准一"), ("目标二", 0.4, "标准二")])
    sheet = parse_offline_objective_sheet(data)
    assert len(sheet.objectives) == 2
    assert any("80" in w for w in sheet.warnings)


# ---------- 接口 1：目标周期导入 ----------


def test_cycle_import_writes_draft_and_auto_joins_participant(client: TestClient) -> None:
    hr_token = _login(client, "mock-hr")
    oc_id = _create_objective_cycle(client, hr_token, "线下导入测试周期")
    bob_id = _user_id("mock-bob")

    data = _make_offline_sheet(
        "mock-bob", "李 Bob",
        [("线下目标一", 0.6, "标准一"), ("线下目标二", 0.4, "标准二")],
    )
    resp = _upload_cycle(client, hr_token, oc_id, [("bob.xlsx", data)])
    assert resp.status_code == 200, resp.text
    result = resp.json()
    assert result["imported_users"] == 1
    assert result["imported_objectives"] == 2
    assert result["skipped"] == []

    with Session(engine) as s:
        objs = s.exec(
            select(Objective)
            .where(Objective.objective_cycle_id == oc_id, Objective.user_id == bob_id)
            .order_by(Objective.order_num)
        ).all()
        assert [o.title for o in objs] == ["线下目标一", "线下目标二"]
        assert all(o.status == "draft" for o in objs)
        assert [o.weight for o in objs] == [60, 40]
        assert objs[0].description == "标准一"
        assert objs[0].measure_criteria == "标准一"

        # 原本不在参与人里，导入后自动加入
        participant = s.exec(
            select(ObjectiveCycleParticipant).where(
                ObjectiveCycleParticipant.objective_cycle_id == oc_id,
                ObjectiveCycleParticipant.user_id == bob_id,
            )
        ).first()
        assert participant is not None


def test_cycle_import_skips_approved_user(client: TestClient) -> None:
    hr_token = _login(client, "mock-hr")
    # seed 周期：active 状态，mock-alice 目标已 approved
    resp = client.get("/api/v1/objective-cycles", headers=_headers(hr_token))
    assert resp.status_code == 200, resp.text
    oc_id = resp.json()[0]["id"]
    alice_id = _user_id("mock-alice")

    data = _make_offline_sheet("mock-alice", "张 Alice", [("新目标", 1.0, "新标准")])
    resp = _upload_cycle(client, hr_token, oc_id, [("alice.xlsx", data)])
    assert resp.status_code == 200, resp.text
    result = resp.json()
    assert result["imported_users"] == 0
    assert len(result["skipped"]) == 1
    assert result["skipped"][0]["wecom_userid"] == "mock-alice"
    assert "确认" in result["skipped"][0]["reason"]

    # 原 approved 目标未被覆盖
    with Session(engine) as s:
        objs = s.exec(
            select(Objective).where(
                Objective.objective_cycle_id == oc_id, Objective.user_id == alice_id
            )
        ).all()
        assert objs and all(o.status == "approved" for o in objs)
        assert "新目标" not in [o.title for o in objs]


def test_cycle_import_skips_unknown_user(client: TestClient) -> None:
    hr_token = _login(client, "mock-hr")
    oc_id = _create_objective_cycle(client, hr_token, "线下导入未知员工周期")
    data = _make_offline_sheet("no-such-user", "不存在", [("目标一", 1.0, "标准一")])
    resp = _upload_cycle(client, hr_token, oc_id, [("ghost.xlsx", data)])
    assert resp.status_code == 200, resp.text
    result = resp.json()
    assert result["imported_users"] == 0
    assert result["skipped"][0]["wecom_userid"] == "no-such-user"
    assert "不存在" in result["skipped"][0]["reason"]


# ---------- 接口 2：试用期目标导入 ----------


def test_probation_import_overwrite(client: TestClient) -> None:
    plan_id = _reset_plan("mock-alice")
    hr_token = _login(client, "mock-hr")

    data = _make_offline_sheet(
        "mock-alice", "张 Alice",
        [("试用目标一", 0.5, "关键成果一"), ("试用目标二", 0.5, "关键成果二")],
    )
    resp = _upload_probation(client, hr_token, [("alice.xlsx", data)])
    assert resp.status_code == 200, resp.text
    result = resp.json()
    assert result["imported_users"] == 1
    assert result["imported_objectives"] == 2
    assert result["skipped"] == []

    with Session(engine) as s:
        objs = s.exec(
            select(ProbationObjective)
            .where(ProbationObjective.plan_id == plan_id)
            .order_by(ProbationObjective.order_num)
        ).all()
        assert [o.title for o in objs] == ["试用目标一", "试用目标二"]
        assert objs[0].description == "关键成果一"
        assert objs[0].measure_criteria == "关键成果一"

    # 再次导入覆盖而不是追加
    data2 = _make_offline_sheet("mock-alice", "张 Alice", [("覆盖目标", 1.0, "覆盖成果")])
    resp = _upload_probation(client, hr_token, [("alice2.xlsx", data2)])
    assert resp.status_code == 200, resp.text
    assert resp.json()["imported_objectives"] == 1
    with Session(engine) as s:
        objs = s.exec(
            select(ProbationObjective).where(ProbationObjective.plan_id == plan_id)
        ).all()
        assert [o.title for o in objs] == ["覆盖目标"]


def test_probation_import_skip_when_plan_not_exists(client: TestClient) -> None:
    # mock-bob 没有试用期计划（先删子表目标再删计划，避免外键拦截）
    with Session(engine) as s:
        bob = s.exec(select(User).where(User.wecom_userid == "mock-bob")).first()
        assert bob
        plans = s.exec(select(ProbationPlan).where(ProbationPlan.user_id == bob.id)).all()
        for p in plans:
            for o in s.exec(select(ProbationObjective).where(ProbationObjective.plan_id == p.id)).all():
                s.delete(o)
            s.delete(p)
        s.commit()

    hr_token = _login(client, "mock-hr")
    data = _make_offline_sheet("mock-bob", "李 Bob", [("目标一", 1.0, "成果一")])
    resp = _upload_probation(client, hr_token, [("bob.xlsx", data)])
    assert resp.status_code == 200, resp.text
    result = resp.json()
    assert result["imported_users"] == 0
    assert len(result["skipped"]) == 1
    assert result["skipped"][0]["wecom_userid"] == "mock-bob"
    assert "不存在" in result["skipped"][0]["reason"]


def test_probation_import_forbidden_for_non_hr(client: TestClient) -> None:
    _reset_plan("mock-alice")
    alice_token = _login(client, "mock-alice")
    data = _make_offline_sheet("mock-alice", "张 Alice", [("目标一", 1.0, "成果一")])
    resp = _upload_probation(client, alice_token, [("alice.xlsx", data)])
    assert resp.status_code == 403, resp.text
