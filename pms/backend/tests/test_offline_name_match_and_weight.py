"""线下导入按姓名匹配 + 试用期目标权重测试。

覆盖：
- 姓名匹配：无工号按姓名导入成功；姓名不存在 skip；重名 skip
- 权重：线下导入试用期目标 weight 落库；save 接口带 weight；查看接口返回 weight
"""

from __future__ import annotations

import io
from datetime import date, timedelta
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlmodel import Session, select

from pms.database.models.enums import ProbationPlanStatus
from pms.database.models.objective import Objective
from pms.database.models.objective_cycle_participant import ObjectiveCycleParticipant
from pms.database.models.probation import ProbationObjective, ProbationPlan
from pms.database.models.user import User
from pms.database.session import engine
from pms.main import app

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
FIXTURE = Path(__file__).parent / "fixtures" / "offline_objective_template.xlsx"


@pytest.fixture
def client():
    return TestClient(app)


def _login(client: TestClient, wecom_userid: str) -> str:
    resp = client.post("/api/v1/auth/mock-login", json={"wecom_userid": wecom_userid})
    assert resp.status_code == 200, resp.text
    return resp.json()["token"]


def _headers(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


def _make_offline_sheet(
    wecom_userid: str,
    name: str,
    objectives: list[tuple[str, object, str]],
) -> bytes:
    """在真实样本模板上填充 姓名/工号/目标行，返回 xlsx 字节。"""
    wb = load_workbook(FIXTURE)
    ws = wb["目标设定表"]
    ws["B2"] = name
    ws["F2"] = wecom_userid
    for i, (title, weight, measure) in enumerate(objectives):
        row = 9 + i
        ws.cell(row=row, column=2, value=title)
        ws.cell(row=row, column=3, value=weight)
        ws.cell(row=row, column=4, value=measure)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload_cycle(
    client: TestClient, token: str, objective_cycle_id: int, files: list[tuple[str, bytes]]
):
    return client.post(
        f"/api/v1/objective-cycles/{objective_cycle_id}/excel/import-offline",
        headers=_headers(token),
        files=[("files", (fname, data, XLSX_MIME)) for fname, data in files],
    )


def _upload_probation(client: TestClient, token: str, files: list[tuple[str, bytes]]):
    return client.post(
        "/api/v1/probation/import-offline-objectives",
        headers=_headers(token),
        files=[("files", (fname, data, XLSX_MIME)) for fname, data in files],
    )


def _create_objective_cycle(client: TestClient, hr_token: str, name: str) -> int:
    resp = client.post(
        "/api/v1/objective-cycles",
        headers=_headers(hr_token),
        json={"name": name, "start_date": "2025-07-01", "end_date": "2025-12-31"},
    )
    assert resp.status_code == 200, resp.text
    return resp.json()["id"]


def _user_id(wecom_userid: str) -> int:
    with Session(engine) as s:
        user = s.exec(select(User).where(User.wecom_userid == wecom_userid)).first()
        assert user, f"seed 用户 {wecom_userid} 不存在"
        return user.id


def _reset_plan(wecom_userid: str, status: str = ProbationPlanStatus.DRAFT) -> int:
    """清空该用户的试用期计划并重建一个指定状态的计划，返回 plan.id。"""
    with Session(engine) as s:
        user = s.exec(select(User).where(User.wecom_userid == wecom_userid)).first()
        assert user, f"seed 用户 {wecom_userid} 不存在"
        old_plans = s.exec(select(ProbationPlan).where(ProbationPlan.user_id == user.id)).all()
        for p in old_plans:
            for o in s.exec(
                select(ProbationObjective).where(ProbationObjective.plan_id == p.id)
            ).all():
                s.delete(o)
            s.delete(p)
        # 同一 flush 内 SQLAlchemy 先 INSERT 后 DELETE，
        # 会撞 uq_probation_plan_user，先 flush 把删除落库再建新计划
        s.flush()
        plan = ProbationPlan(
            user_id=user.id,
            start_date=date.today() - timedelta(days=30),
            end_date=date.today() + timedelta(days=150),
            probation_months=6,
            status=status,
        )
        s.add(plan)
        s.commit()
        s.refresh(plan)
        return plan.id


# ---------- 姓名匹配：目标周期导入 ----------


def test_cycle_import_by_name_success(client: TestClient) -> None:
    # 工号留空，按姓名匹配到唯一 active 用户
    hr_token = _login(client, "mock-hr")
    oc_id = _create_objective_cycle(client, hr_token, "姓名匹配导入周期")
    bob_id = _user_id("mock-bob")

    data = _make_offline_sheet(
        "", "李 Bob",
        [("姓名匹配目标一", 0.6, "标准一"), ("姓名匹配目标二", 0.4, "标准二")],
    )
    resp = _upload_cycle(client, hr_token, oc_id, [("bob.xlsx", data)])
    assert resp.status_code == 200, resp.text
    result = resp.json()
    assert result["imported_users"] == 1
    assert result["imported_objectives"] == 2
    assert result["skipped"] == []

    with Session(engine) as s:
        objs = s.exec(
            select(Objective)
            .where(Objective.objective_cycle_id == oc_id, Objective.user_id == bob_id)
            .order_by(Objective.order_num)
        ).all()
        assert [o.title for o in objs] == ["姓名匹配目标一", "姓名匹配目标二"]


def test_cycle_import_by_name_not_found(client: TestClient) -> None:
    hr_token = _login(client, "mock-hr")
    oc_id = _create_objective_cycle(client, hr_token, "姓名未匹配导入周期")

    data = _make_offline_sheet("", "不存在的人", [("目标一", 1.0, "标准一")])
    resp = _upload_cycle(client, hr_token, oc_id, [("ghost.xlsx", data)])
    assert resp.status_code == 200, resp.text
    result = resp.json()
    assert result["imported_users"] == 0
    assert len(result["skipped"]) == 1
    assert result["skipped"][0]["reason"] == "姓名未匹配"


def test_cycle_import_by_name_duplicate(client: TestClient) -> None:
    # 构造两个同名 active 用户 → skip 提示补充工号
    with Session(engine) as s:
        s.add(User(wecom_userid="mock-dup-1", name="重名用户", status="active"))
        s.add(User(wecom_userid="mock-dup-2", name="重名用户", status="active"))
        s.commit()
    try:
        hr_token = _login(client, "mock-hr")
        oc_id = _create_objective_cycle(client, hr_token, "姓名重名导入周期")
        data = _make_offline_sheet("", "重名用户", [("目标一", 1.0, "标准一")])
        resp = _upload_cycle(client, hr_token, oc_id, [("dup.xlsx", data)])
        assert resp.status_code == 200, resp.text
        result = resp.json()
        assert result["imported_users"] == 0
        assert len(result["skipped"]) == 1
        assert result["skipped"][0]["reason"] == "姓名重名，请在表格中补充工号"

        # 补充工号后按工号匹配成功
        data2 = _make_offline_sheet("mock-dup-1", "重名用户", [("目标一", 1.0, "标准一")])
        resp = _upload_cycle(client, hr_token, oc_id, [("dup2.xlsx", data2)])
        assert resp.status_code == 200, resp.text
        assert resp.json()["imported_users"] == 1
    finally:
        with Session(engine) as s:
            dup_ids = s.exec(
                select(User.id).where(User.wecom_userid.in_(("mock-dup-1", "mock-dup-2")))
            ).all()
            if dup_ids:
                for o in s.exec(select(Objective).where(Objective.user_id.in_(dup_ids))).all():
                    s.delete(o)
                for p in s.exec(
                    select(ObjectiveCycleParticipant).where(
                        ObjectiveCycleParticipant.user_id.in_(dup_ids)
                    )
                ).all():
                    s.delete(p)
            for uid in ("mock-dup-1", "mock-dup-2"):
                user = s.exec(select(User).where(User.wecom_userid == uid)).first()
                if user:
                    s.delete(user)
            s.commit()


def test_probation_import_by_name_not_found(client: TestClient) -> None:
    hr_token = _login(client, "mock-hr")
    data = _make_offline_sheet("", "不存在的人", [("目标一", 1.0, "成果一")])
    resp = _upload_probation(client, hr_token, [("ghost.xlsx", data)])
    assert resp.status_code == 200, resp.text
    result = resp.json()
    assert result["imported_users"] == 0
    assert len(result["skipped"]) == 1
    assert result["skipped"][0]["reason"] == "姓名未匹配"


# ---------- 权重：线下导入落库 ----------


def test_probation_offline_import_writes_weight(client: TestClient) -> None:
    # 工号留空按姓名匹配，且解析出的 weight 落库（此前被丢弃）
    plan_id = _reset_plan("mock-bob")
    hr_token = _login(client, "mock-hr")

    data = _make_offline_sheet(
        "", "李 Bob",
        [("试用目标一", 0.6, "关键成果一"), ("试用目标二", 0.4, "关键成果二")],
    )
    resp = _upload_probation(client, hr_token, [("bob.xlsx", data)])
    assert resp.status_code == 200, resp.text
    result = resp.json()
    assert result["imported_users"] == 1
    assert result["imported_objectives"] == 2
    assert result["skipped"] == []

    with Session(engine) as s:
        objs = s.exec(
            select(ProbationObjective)
            .where(ProbationObjective.plan_id == plan_id)
            .order_by(ProbationObjective.order_num)
        ).all()
        assert [o.title for o in objs] == ["试用目标一", "试用目标二"]
        assert [o.weight for o in objs] == [60, 40]


# ---------- 权重：save / 查看接口 ----------


def test_probation_save_and_view_weight(client: TestClient) -> None:
    _reset_plan("mock-alice")
    alice_token = _login(client, "mock-alice")
    alice_id = _user_id("mock-alice")

    resp = client.post(
        f"/api/v1/probation/{alice_id}/objectives",
        headers=_headers(alice_token),
        json={
            "objectives": [
                {"title": "目标一", "description": "描述一", "measure_criteria": "标准一", "weight": 70},
                {"title": "目标二", "description": "描述二", "measure_criteria": "标准二", "weight": 30},
            ],
            "submit": False,
        },
    )
    assert resp.status_code == 200, resp.text
    assert resp.json()["saved"] == 2

    # 查看接口返回 weight
    resp = client.get("/api/v1/probation/mine", headers=_headers(alice_token))
    assert resp.status_code == 200, resp.text
    plan = resp.json()
    assert plan is not None
    assert [o["title"] for o in plan["objectives"]] == ["目标一", "目标二"]
    assert [o["weight"] for o in plan["objectives"]] == [70, 30]
